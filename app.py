import os
import io
import csv
import sqlite3
from datetime import datetime, date
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, g

try:
    import psycopg2
    import psycopg2.extras
    PSYCOPG2_AVAILABLE = True
except ImportError:
    PSYCOPG2_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "logistica_eventos_secret_2026")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "eventos.db")
DATABASE_URL = os.environ.get("DATABASE_URL", "")

# Render entrega postgres://, psycopg2 necesita postgresql://
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

USE_POSTGRES = bool(DATABASE_URL and PSYCOPG2_AVAILABLE)
PH = "%s" if USE_POSTGRES else "?"

TIPOS_EVENTO = {
    "autoelevador": "Autoelevador",
    "grua": "Grua",
    "camion": "Camion",
    "remitero": "Remitero",
    "otro": "Otro",
}
TIPOS_GRUA = ["Portuaria", "Movil", "Portico", "Flotante", "Otra"]

# ── Conexion DB ────────────────────────────────────────────────────────────────

def get_db():
    if "db" not in g:
        if USE_POSTGRES:
            conn = psycopg2.connect(DATABASE_URL)
            conn.autocommit = False
            g.db = conn
        else:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            g.db = conn
    return g.db

def db_execute(sql, params=()):
    conn = get_db()
    if USE_POSTGRES:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params)
        return cur
    else:
        return conn.execute(sql, params)

def db_commit():
    get_db().commit()

@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()

def init_db():
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS operacion (
                id SERIAL PRIMARY KEY,
                nombre TEXT NOT NULL,
                descripcion TEXT,
                fecha_inicio TEXT NOT NULL,
                activa INTEGER DEFAULT 1
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS evento (
                id SERIAL PRIMARY KEY,
                operacion_id INTEGER NOT NULL REFERENCES operacion(id),
                tipo TEXT NOT NULL,
                fecha TEXT NOT NULL,
                hora_inicio TEXT,
                hora_fin TEXT,
                cantidad INTEGER,
                tipo_grua TEXT,
                ubicacion TEXT,
                usuario TEXT,
                observaciones TEXT,
                created_at TEXT NOT NULL
            )
        """)
        conn.commit()
        conn.close()
    else:
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS operacion (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombre TEXT NOT NULL,
                    descripcion TEXT,
                    fecha_inicio TEXT NOT NULL,
                    activa INTEGER DEFAULT 1
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS evento (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    operacion_id INTEGER NOT NULL,
                    tipo TEXT NOT NULL,
                    fecha TEXT NOT NULL,
                    hora_inicio TEXT,
                    hora_fin TEXT,
                    cantidad INTEGER,
                    tipo_grua TEXT,
                    ubicacion TEXT,
                    usuario TEXT,
                    observaciones TEXT,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY (operacion_id) REFERENCES operacion(id)
                )
            """)
            conn.commit()

# ── Rutas ──────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    cur = db_execute(
        "SELECT * FROM operacion WHERE activa = 1 ORDER BY id DESC LIMIT 1"
    )
    operacion = cur.fetchone()
    return render_template("index.html", operacion=operacion)


@app.route("/operacion/nueva", methods=["GET", "POST"])
def nueva_operacion():
    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip()
        descripcion = request.form.get("descripcion", "").strip()
        if not nombre:
            flash("El nombre de la operacion es obligatorio.", "danger")
            return redirect(url_for("nueva_operacion"))
        db_execute("UPDATE operacion SET activa = 0")
        db_execute(
            f"INSERT INTO operacion (nombre, descripcion, fecha_inicio, activa) VALUES ({PH},{PH},{PH},1)",
            (nombre, descripcion, date.today().isoformat()),
        )
        db_commit()
        flash(f"Operacion '{nombre}' iniciada.", "success")
        return redirect(url_for("index"))
    return render_template("nueva_operacion.html")


@app.route("/evento/nuevo/<tipo>", methods=["GET", "POST"])
def nuevo_evento(tipo):
    if tipo not in TIPOS_EVENTO:
        flash("Tipo de evento invalido.", "danger")
        return redirect(url_for("index"))

    cur = db_execute(
        "SELECT * FROM operacion WHERE activa = 1 ORDER BY id DESC LIMIT 1"
    )
    operacion = cur.fetchone()
    if not operacion:
        flash("Primero crea una operacion activa.", "warning")
        return redirect(url_for("nueva_operacion"))

    if request.method == "POST":
        fecha = request.form.get("fecha") or date.today().isoformat()
        hora_inicio = request.form.get("hora_inicio") or None
        hora_fin = request.form.get("hora_fin") or None
        cantidad_raw = request.form.get("cantidad") or None
        cantidad = int(cantidad_raw) if cantidad_raw and cantidad_raw.isdigit() else None
        tipo_grua = request.form.get("tipo_grua") or None
        ubicacion = request.form.get("ubicacion", "").strip() or None
        usuario = request.form.get("usuario", "").strip() or None
        observaciones = request.form.get("observaciones", "").strip() or None

        db_execute(
            f"""INSERT INTO evento
               (operacion_id, tipo, fecha, hora_inicio, hora_fin, cantidad,
                tipo_grua, ubicacion, usuario, observaciones, created_at)
               VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH})""",
            (
                operacion["id"], tipo, fecha, hora_inicio, hora_fin, cantidad,
                tipo_grua, ubicacion, usuario, observaciones,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        db_commit()
        flash(f"{TIPOS_EVENTO[tipo]} registrado.", "success")
        return redirect(url_for("index"))

    return render_template(
        "nuevo_evento.html",
        tipo=tipo,
        tipo_label=TIPOS_EVENTO[tipo],
        operacion=operacion,
        tipos_grua=TIPOS_GRUA,
        hoy=date.today().isoformat(),
    )


@app.route("/resumen")
def resumen():
    cur = db_execute(
        "SELECT * FROM operacion WHERE activa = 1 ORDER BY id DESC LIMIT 1"
    )
    operacion = cur.fetchone()
    if not operacion:
        flash("No hay operacion activa.", "warning")
        return redirect(url_for("index"))

    eventos_por_tipo = {}
    for tipo, label in TIPOS_EVENTO.items():
        rows = db_execute(
            f"SELECT * FROM evento WHERE operacion_id = {PH} AND tipo = {PH} ORDER BY fecha, hora_inicio",
            (operacion["id"], tipo),
        ).fetchall()
        eventos_por_tipo[tipo] = {"label": label, "eventos": rows}

    total = db_execute(
        f"SELECT COUNT(*) as n FROM evento WHERE operacion_id = {PH}",
        (operacion["id"],),
    ).fetchone()["n"]

    return render_template(
        "resumen.html",
        operacion=operacion,
        eventos_por_tipo=eventos_por_tipo,
        tipos_evento=TIPOS_EVENTO,
        total=total,
    )


@app.route("/evento/eliminar/<int:evento_id>", methods=["POST"])
def eliminar_evento(evento_id):
    db_execute(f"DELETE FROM evento WHERE id = {PH}", (evento_id,))
    db_commit()
    flash("Evento eliminado.", "info")
    return redirect(url_for("resumen"))


# ── Descarga ───────────────────────────────────────────────────────────────────

def _get_eventos(operacion_id):
    return db_execute(
        f"""SELECT tipo, fecha, hora_inicio, hora_fin, cantidad,
                  tipo_grua, ubicacion, usuario, observaciones, created_at
           FROM evento WHERE operacion_id = {PH}
           ORDER BY tipo, fecha, hora_inicio""",
        (operacion_id,),
    ).fetchall()


@app.route("/descargar/csv")
def descargar_csv():
    cur = db_execute(
        "SELECT * FROM operacion WHERE activa = 1 ORDER BY id DESC LIMIT 1"
    )
    operacion = cur.fetchone()
    if not operacion:
        flash("No hay operacion activa.", "warning")
        return redirect(url_for("index"))

    eventos = _get_eventos(operacion["id"])
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Operacion", operacion["nombre"]])
    writer.writerow(["Inicio", operacion["fecha_inicio"]])
    writer.writerow([])
    writer.writerow([
        "Tipo", "Fecha", "Hora Inicio", "Hora Fin", "Cantidad",
        "Tipo Grua", "Ubicacion/Descripcion", "Usuario", "Observaciones", "Registrado"
    ])
    for e in eventos:
        writer.writerow([
            TIPOS_EVENTO.get(e["tipo"], e["tipo"]),
            e["fecha"], e["hora_inicio"] or "", e["hora_fin"] or "",
            e["cantidad"] or "", e["tipo_grua"] or "", e["ubicacion"] or "",
            e["usuario"] or "", e["observaciones"] or "", e["created_at"],
        ])

    output.seek(0)
    nombre_archivo = f"operacion_{operacion['nombre'].replace(' ', '_')}_{date.today().isoformat()}.csv"
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=nombre_archivo,
    )


@app.route("/descargar/excel")
def descargar_excel():
    if not EXCEL_AVAILABLE:
        flash("openpyxl no esta instalado.", "warning")
        return redirect(url_for("resumen"))

    cur = db_execute(
        "SELECT * FROM operacion WHERE activa = 1 ORDER BY id DESC LIMIT 1"
    )
    operacion = cur.fetchone()
    if not operacion:
        flash("No hay operacion activa.", "warning")
        return redirect(url_for("index"))

    eventos = _get_eventos(operacion["id"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eventos"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Operacion: {operacion['nombre']}  |  Inicio: {operacion['fecha_inicio']}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    col_headers = [
        "Tipo", "Fecha", "Hora Inicio", "Hora Fin", "Cantidad",
        "Tipo Grua", "Ubicacion/Descripcion", "Usuario", "Observaciones", "Registrado"
    ]
    col_fill = PatternFill("solid", fgColor="2E75B6")
    col_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = col_fill
        cell.font = col_font

    alt_fill = PatternFill("solid", fgColor="DDEEFF")
    for row_idx, e in enumerate(eventos, start=3):
        fill = alt_fill if row_idx % 2 == 0 else None
        values = [
            TIPOS_EVENTO.get(e["tipo"], e["tipo"]),
            e["fecha"], e["hora_inicio"] or "", e["hora_fin"] or "",
            e["cantidad"] or "", e["tipo_grua"] or "", e["ubicacion"] or "",
            e["usuario"] or "", e["observaciones"] or "", e["created_at"],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fill:
                cell.fill = fill

    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Tipo"
    ws2["B1"] = "Cantidad de registros"
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    for row_idx, (tipo, label) in enumerate(TIPOS_EVENTO.items(), start=2):
        count = sum(1 for e in eventos if e["tipo"] == tipo)
        ws2.cell(row=row_idx, column=1, value=label)
        ws2.cell(row=row_idx, column=2, value=count)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre_archivo = f"operacion_{operacion['nombre'].replace(' ', '_')}_{date.today().isoformat()}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_archivo,
    )


# ── Main ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_db()
    import socket
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = "127.0.0.1"
    print(f"\n{'='*50}")
    print(f"  App de Eventos Logisticos")
    print(f"  Local:  http://127.0.0.1:5000")
    print(f"  Red:    http://{local_ip}:5000")
    print(f"  BD:     {'PostgreSQL' if USE_POSTGRES else 'SQLite (local)'}")
    print(f"{'='*50}\n")
    app.run(host="0.0.0.0", port=5000, debug=False)
